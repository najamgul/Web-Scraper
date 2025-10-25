# routes.py
# routes.py
import os
import re
import json
import csv
import io
import logging
import hashlib
import math
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError
from datetime import datetime, timedelta
import time  # Should already be there
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError

from app.enrichment import enrich_threat_intelligence
from flask import session, Blueprint, render_template, request, jsonify, send_file, redirect, url_for, flash
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash
from bson import ObjectId
from app.analytics import (
    get_dashboard_stats,
    get_top_malicious_ips,
    get_top_malicious_domains,
    get_geolocation_data,
    get_threat_timeline,
    get_recent_threats,
    get_classification_breakdown,
    get_top_threat_tags
)

from app.scraper import google_cse_search
from app.vt_shodan_api import vt_lookup_domain, vt_lookup_ip, vt_lookup_url, shodan_lookup
from app.otx_api import otx_lookup

# ✅ Import Unified Orchestrator
from app.orchestrator import (
    orchestrate_threat_intelligence,
    detect_input_type as orchestrator_detect_type,
    format_results_for_template as orchestrator_format_results
)

# ✅ INITIALIZE LOGGER
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Use improved ML model
try:
    from app.ml_model_improved import classify_threat, classify_threat_with_details
    logger.info("✅ Using improved ML model with TF-IDF and SMOTE")
except ImportError:
    from app.ml_model import classify_threat
    classify_threat_with_details = None
    logger.warning("⚠️ Falling back to legacy ML model")

from app.forms import InputForm, LoginForm, SignupForm
from app.models import User, IOCResult, Feedback
from app.decorators import login_required

from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

try:
    import numpy as np
    NUMPY_INT_TYPES = (np.integer,)
except ImportError:  # numpy is optional at runtime
    NUMPY_INT_TYPES = tuple()

# ✅ CACHE SETUP
CACHE = {}
CACHE_DURATION = timedelta(hours=1)

def get_cached_result(ioc_value):
    """Check cache for recent scan"""
    cache_key = hashlib.md5(ioc_value.encode()).hexdigest()
    if cache_key in CACHE:
        cached_data, timestamp = CACHE[cache_key]
        if datetime.utcnow() - timestamp < CACHE_DURATION:
            logger.info(f"✅ Cache HIT for {ioc_value}")
            return cached_data
    return None

def set_cached_result(ioc_value, data):
    """Cache scan result"""
    cache_key = hashlib.md5(ioc_value.encode()).hexdigest()
    CACHE[cache_key] = (data, datetime.utcnow())
    logger.info(f"💾 Cached result for {ioc_value}")


MAX_MONGO_INT = (1 << 63) - 1


def _sanitize_scalar(value):
    """Normalize scalars so MongoDB can store them safely."""
    if isinstance(value, bool):  # bool is a subclass of int, keep as-is
        return value

    if isinstance(value, NUMPY_INT_TYPES):
        value = int(value)

    if isinstance(value, int):
        return str(value) if abs(value) > MAX_MONGO_INT else value

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return str(value)
        return value

    return value


def sanitize_for_mongo(data):
    """Recursively convert values that MongoDB cannot store (e.g. >8-byte ints)."""
    if isinstance(data, dict):
        return {key: sanitize_for_mongo(value) for key, value in data.items()}
    if isinstance(data, list):
        return [sanitize_for_mongo(item) for item in data]
    return _sanitize_scalar(data)


# ✅ PARALLEL API FETCHER
def fetch_all_threat_data(user_input, ioc_type):
    """
    Fetch all API data in PARALLEL with detailed logging
    """
    results = {
        'vt_data': {},
        'shodan_data': {},
        'otx_data': {},
        'scraped_data': []
    }
    
    logger.info(f"📡 Fetching threat data for {ioc_type}: {user_input}")
    
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {}
        
        # Submit all API calls SIMULTANEOUSLY
        logger.info("   Submitting OTX lookup...")
        futures['otx'] = executor.submit(otx_lookup, user_input, ioc_type)
        
        if ioc_type == "ip":
            logger.info("   Submitting VT IP lookup...")
            futures['vt'] = executor.submit(vt_lookup_ip, user_input)
            logger.info("   Submitting Shodan lookup...")
            futures['shodan'] = executor.submit(shodan_lookup, user_input)
        elif ioc_type == "url":
            logger.info("   Submitting VT URL lookup...")
            futures['vt'] = executor.submit(vt_lookup_url, user_input)
        elif ioc_type in ["domain", "hash"]:
            logger.info("   Submitting VT domain/hash lookup...")
            futures['vt'] = executor.submit(vt_lookup_domain, user_input)
        
        # Only scrape for keywords
        if ioc_type == "keyword":
            logger.info("   Submitting Google CSE search...")
            futures['scraped'] = executor.submit(google_cse_search, user_input)
        
        # Collect results with timeout protection
        for key, future in futures.items():
            try:
                if key == 'vt':
                    result = future.result(timeout=15)  # Increased timeout
                    results['vt_data'] = result or {}
                    logger.info(f"   ✅ VT returned: {len(str(result))} chars")
                    if 'error' in results['vt_data']:
                        logger.warning(f"   ⚠️ VT error: {results['vt_data']['error']}")
                    
                elif key == 'shodan':
                    result = future.result(timeout=15)
                    results['shodan_data'] = result or {}
                    logger.info(f"   ✅ Shodan returned: {len(str(result))} chars")
                    if 'error' in results['shodan_data']:
                        logger.warning(f"   ⚠️ Shodan error: {results['shodan_data']['error']}")
                    
                elif key == 'otx':
                    result = future.result(timeout=20)
                    results['otx_data'] = result or {}
                    logger.info(f"   ✅ OTX returned: {len(str(result))} chars, classification: {result.get('classification', 'N/A')}")
                    if 'error' in results['otx_data']:
                        logger.warning(f"   ⚠️ OTX error: {results['otx_data']['error']}")
                    
                elif key == 'scraped':
                    result = future.result(timeout=20)
                    results['scraped_data'] = result or []
                    logger.info(f"   ✅ Scraped {len(results['scraped_data'])} results")
                    
            except TimeoutError:
                logger.error(f"   ❌ {key} API TIMEOUT after waiting")
            except Exception as e:
                logger.error(f"   ❌ {key} API ERROR: {e}", exc_info=True)
    
    # Log final results summary
    logger.info(f"📊 API Summary:")
    logger.info(f"   VT data: {'✅ OK' if results['vt_data'] and 'error' not in results['vt_data'] else '❌ Failed'}")
    logger.info(f"   Shodan data: {'✅ OK' if results['shodan_data'] and 'error' not in results['shodan_data'] else '❌ Failed'}")
    logger.info(f"   OTX data: {'✅ OK' if results['otx_data'] and 'error' not in results['otx_data'] else '❌ Failed'}")
    
    return results


# Add to routes.py after imports
def format_otx_for_display(otx_data):
    """
    Format OTX data for template display
    """
    if not otx_data or 'error' in otx_data:
        return None
    
    details = otx_data.get('details', {})
    
    # Extract key information
    formatted = {
        'threat_score': otx_data.get('threat_score', 0),
        'classification': otx_data.get('classification', 'Unknown'),
        'source_count': otx_data.get('source_count', 0),
        'summary': details.get('summary', 'No summary available'),
        'risk_level': details.get('risk_level', 'Unknown'),
        
        # Lists and counts
        'pulses': details.get('pulses', 0),
        'malicious_indicators': details.get('malicious_indicators', 0),
        'malicious_reports': details.get('malicious_reports', 0),
        
        # Arrays
        'malware_families': details.get('malware_families', []),
        'attack_techniques': details.get('attack_techniques', []),
        'top_tags': details.get('top_tags', []),
        'adversaries': details.get('adversaries', []),
        'top_pulses': details.get('top_pulses', []),
        
        # Location (for IPs)
        'country': details.get('country', ''),
        'city': details.get('city', ''),
        'asn': details.get('asn', ''),
        
        # Other
        'reputation': details.get('reputation', 0),
        'vulnerabilities': details.get('vulnerabilities', 0),
        
        # Raw data
        'raw': otx_data
    }
    
    return formatted

def format_google_for_display(scraped_data):
    """
    Format Google Custom Search results for template display
    Similar to OTX formatting
    """
    if not scraped_data or not isinstance(scraped_data, list):
        return None
    
    # Filter out "No scraped data found" entries
    valid_results = [
        item for item in scraped_data 
        if isinstance(item, dict) and item.get('title', '').lower() != 'no scraped data found'
    ]
    
    if not valid_results:
        return None
    
    # Extract domains
    domains = []
    for result in valid_results:
        domain = result.get('displayLink', '')
        if domain:
            domains.append(domain)
    
    unique_domains = list(set(domains))
    
    # Analyze content for threat indicators
    threat_keywords = [
        'malware', 'virus', 'trojan', 'ransomware', 'phishing', 'scam', 
        'fraud', 'hack', 'exploit', 'vulnerability', 'breach', 'attack',
        'suspicious', 'malicious', 'threat', 'compromise', 'botnet', 'backdoor'
    ]
    
    safe_keywords = [
        'legitimate', 'safe', 'secure', 'trusted', 'official', 'verified',
        'protection', 'antivirus', 'security', 'defense'
    ]
    
    threat_indicators = []
    safe_indicators = []
    content_summary = []
    
    for result in valid_results:
        content = (result.get('snippet', '') + ' ' + result.get('content', '')).lower()
        title = result.get('title', '').lower()
        
        # Check for threat keywords
        for keyword in threat_keywords:
            if keyword in content or keyword in title:
                if keyword not in threat_indicators:
                    threat_indicators.append(keyword)
        
        # Check for safe keywords
        for keyword in safe_keywords:
            if keyword in content or keyword in title:
                if keyword not in safe_indicators:
                    safe_indicators.append(keyword)
        
        # Extract first meaningful sentence from snippet
        snippet = result.get('snippet', '')
        if snippet and len(snippet) > 20:
            content_summary.append(snippet[:150] + '...' if len(snippet) > 150 else snippet)
    
    # Calculate threat score based on findings
    threat_score = min(len(threat_indicators) * 10, 100)
    
    # Adjust based on safe indicators
    if safe_indicators:
        threat_score = max(threat_score - len(safe_indicators) * 5, 0)
    
    # Determine classification
    if threat_score >= 60:
        classification = "High Risk"
        risk_level = "Critical"
    elif threat_score >= 30:
        classification = "Medium Risk"
        risk_level = "High"
    elif threat_score > 0:
        classification = "Low Risk"
        risk_level = "Medium"
    else:
        classification = "Informational"
        risk_level = "Low"
    
    # Build summary
    if threat_indicators:
        summary = f"Found {len(valid_results)} search results with {len(threat_indicators)} threat-related indicators across {len(unique_domains)} domains"
    else:
        summary = f"Found {len(valid_results)} informational search results across {len(unique_domains)} domains with no immediate threat indicators"
    
    formatted = {
        'result_count': len(valid_results),
        'threat_score': threat_score,
        'classification': classification,
        'risk_level': risk_level,
        'summary': summary,
        
        # Domain analysis
        'total_domains': len(unique_domains),
        'unique_domains': unique_domains[:10],  # Top 10 domains
        
        # Content analysis
        'threat_indicators': threat_indicators[:10],
        'safe_indicators': safe_indicators[:5],
        'content_previews': content_summary[:5],
        
        # Top results
        'top_results': [
            {
                'title': r.get('title', 'No title'),
                'url': r.get('url', '#'),
                'domain': r.get('displayLink', 'Unknown'),
                'snippet': r.get('snippet', 'No description')[:200]
            }
            for r in valid_results[:5]
        ],
        
        # Raw data
        'raw': scraped_data
    }
    
    return formatted

def classify_keyword_from_google(google_formatted):
    """
    Fallback classification for keywords based on Google search results
    Used when OTX times out or returns no data
    """
    if not google_formatted:
        return "Informational"
    
    threat_score = google_formatted.get('threat_score', 0)
    threat_indicators = google_formatted.get('threat_indicators', [])
    
    # Use Google threat score as fallback
    if threat_score >= 60 or len(threat_indicators) >= 5:
        return "Suspicious"  # Conservative - need OTX confirmation for "Malicious"
    elif threat_score >= 30 or len(threat_indicators) >= 3:
        return "Informational"
    else:
        return "Benign"


main_bp = Blueprint("main", __name__)


def detect_ioc_type(ioc):
    """Detect the type of IOC: IP, URL, domain, hash, or keyword."""
    ip_pattern = r"^(?:\d{1,3}\.){3}\d{1,3}$"
    url_pattern = r"^https?://[^\s/$.?#].[^\s]*$"
    domain_pattern = r"^(?!-)[A-Za-z0-9-]{1,63}(?<!-)(?:\.[A-Za-z]{2,})+$"
    hash_pattern = r"^[a-fA-F0-9]{32,128}$"

    if re.match(ip_pattern, ioc):
        return "ip"
    elif re.match(url_pattern, ioc):
        return "url"
    elif re.match(domain_pattern, ioc):
        return "domain"
    elif re.match(hash_pattern, ioc):
        return "hash"
    else:
        return "keyword"


# ---- LANDING PAGE ----
@main_bp.route("/")
def landing():
    """Public landing page"""
    if 'user_id' in session:
        return redirect(url_for('main.index'))
    return render_template("landing.html")

# In routes.py, add this function
@main_bp.route("/clear-cache")
@login_required
def clear_cache():
    """Clear the scan result cache"""
    global CACHE
    CACHE.clear()
    flash("Cache cleared successfully!", "success")
    return redirect(url_for("main.index"))
# ---- INDEX PAGE (OPTIMIZED) ----
# ---- INDEX PAGE (OPTIMIZED) ----
@main_bp.route("/index", methods=["GET", "POST"])
@login_required
def index():
    form = InputForm()
    if form.validate_on_submit():
        user_input = form.input_data.data.strip()
        if not user_input:
            flash("Please enter a keyword, IP, URL, domain, or hash.", "warning")
            return redirect(url_for("main.index"))

        ioc_type = detect_ioc_type(user_input)
        
        # ✅ LOG SCAN REQUEST
        logger.info(f"\n{'='*80}")
        logger.info(f"📡 SCAN REQUEST (via /index)")
        logger.info(f"   User: {session.get('username', 'Anonymous')}")
        logger.info(f"   Input: {user_input}")
        logger.info(f"   Type: {ioc_type}")
        logger.info(f"{'='*80}\n")
        
        # ✅ CHECK CACHE FIRST
        cached = get_cached_result(user_input)
        if cached:
            logger.info("💾 Returning cached results")
            flash("Results loaded from cache (scanned recently)", "info")
            return render_template("results.html", results=cached['results'], chart_data=cached['chart_data'])
        
        # ✅ FETCH ALL APIs IN PARALLEL
        logger.info(f"🚀 Starting PARALLEL API fetch for {ioc_type}: {user_input}")
        start_time = datetime.utcnow()
        
        
        # ✅ OPTIMIZED: Skip OTX on initial load, fetch via AJAX
        logger.info(f"🚀 Starting FAST API fetch (VT + Shodan + AbuseIPDB)")
        start_time = datetime.utcnow()

# Fetch only VT, Shodan, AbuseIPDB and Google (skip OTX)
        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = {}

            # VirusTotal
            if ioc_type == "ip":
                logger.info("   📌 Submitting VT IP lookup...")
                futures['vt'] = executor.submit(vt_lookup_ip, user_input)
            elif ioc_type == "url":
                logger.info("   📌 Submitting VT URL lookup...")
                futures['vt'] = executor.submit(vt_lookup_url, user_input)
            elif ioc_type in ["domain", "hash"]:
                logger.info("   📌 Submitting VT domain/hash lookup...")
                futures['vt'] = executor.submit(vt_lookup_domain, user_input)

            # Shodan (only for IPs)
            if ioc_type == "ip":
                logger.info("   📌 Submitting Shodan lookup...")
                futures['shodan'] = executor.submit(shodan_lookup, user_input)
                
            # AbuseIPDB (only for IPs)
            if ioc_type == "ip":
                logger.info("   📌 Submitting AbuseIPDB lookup...")
                from app.abuseipdb_api import abuseipdb_lookup
                futures['abuseipdb'] = executor.submit(abuseipdb_lookup, user_input)

            # Google (only for keywords)
            if ioc_type == "keyword":
                logger.info("   📌 Submitting Google CSE search...")
                futures['google'] = executor.submit(google_cse_search, user_input)

            # Collect results with timeout
            vt_data = {}
            shodan_data = {}
            abuseipdb_data = {}
            scraped_data = []

            for key, future in futures.items():
                try:
                    if key == 'vt':
                        vt_data = future.result(timeout=15) or {}
                        logger.info(f"   ✅ VT result: {len(str(vt_data))} chars")
                    elif key == 'shodan':
                        shodan_data = future.result(timeout=15) or {}
                        logger.info(f"   ✅ Shodan result: {len(str(shodan_data))} chars")
                    elif key == 'abuseipdb':
                        abuseipdb_data = future.result(timeout=15) or {}
                        logger.info(f"   ✅ AbuseIPDB result: {len(str(abuseipdb_data))} chars")
                    elif key == 'google':
                        scraped_data = future.result(timeout=20) or []
                        logger.info(f"   ✅ Google CSE: {len(scraped_data)} results")
                except Exception as e:
                    logger.error(f"   ❌ Error fetching {key}: {e}")

# ✅ OTX will be loaded via AJAX
        otx_data = {'loading': True}  # Placeholder

        elapsed = (datetime.utcnow() - start_time).total_seconds()
        logger.info(f"⏱️ Fast API fetch completed in {elapsed:.2f} seconds (OTX will load via AJAX)")
        logger.info(f"\n📊 API Results Summary:")
        logger.info(f"   VT: {'✅ OK' if vt_data and 'error' not in vt_data else '❌ Failed'}")
        logger.info(f"   Shodan: {'✅ OK' if shodan_data and 'error' not in shodan_data else '❌ Failed'}")
        logger.info(f"   AbuseIPDB: {'✅ OK' if abuseipdb_data and 'error' not in abuseipdb_data else '❌ Failed'}")
        logger.info(f"   Google CSE: {'✅ OK' if scraped_data else '❌ Failed'}\n")
        
    
        
        # ML Classification (fast)
        # ML Classification
# ✅    # ✅ SKIP ML classification on initial load - will be done via AJAX after OTX loads
        classification = None # Placeholder
        
        # ✅ CRITICAL: Skip AI enrichment entirely on page load
        # It will be loaded via AJAX after results page renders
        enrichment = None  # ← Don't even set a placeholder
        
        # Save to MongoDB
        logger.info("💾 Saving results to MongoDB...")
        user_ref = None
        if session.get("user_id"):
            try:
                user_ref = User.objects.get(id=session.get("user_id"))
            except Exception:
                pass

        vt_for_db = sanitize_for_mongo(vt_data)
        shodan_for_db = sanitize_for_mongo(shodan_data)
        abuseipdb_for_db = sanitize_for_mongo(abuseipdb_data)
        scraped_for_db = sanitize_for_mongo(scraped_data)

        ioc_result = IOCResult(
            input_value=user_input,
            type=ioc_type,
            vt_report=vt_for_db,
            shodan_report=shodan_for_db,
            abuseipdb_report=abuseipdb_for_db,
            otx_report={'loading': True},
            scraped_data=scraped_for_db,
            classification=None,  # ← Will be populated by AJAX
            enrichment_context=None,  # ← Will be populated by AJAX
            user_id=user_ref,
            timestamp=datetime.utcnow()
        )
        ioc_result.save()
        logger.info(f"✅ Saved to MongoDB with ID: {ioc_result.id}")

        # Chart Data
        chart_data = {
            "labels": ["Malicious", "Benign", "Informational", "Unknown"],
            "values": [
                IOCResult.objects(classification="Malicious").count(),
                IOCResult.objects(classification="Benign").count(),
                IOCResult.objects(classification="Informational").count(),
                IOCResult.objects(classification="Unknown").count(),
            ]
        }
        
        results = {
            "ioc_id": str(ioc_result.id),
            "input": user_input,
            "type": ioc_type,
            "vt": vt_data,
            "shodan": shodan_data,
            "abuseipdb": abuseipdb_data,
            "otx": otx_data,  # Will be {'loading': True}
            "otx_formatted": None,  # Will be loaded via AJAX
            "scraped": scraped_data,
            "google_formatted": format_google_for_display(scraped_data) if scraped_data else None,
            "classification": classification,  # Will be "Loading..."
            "enrichment": None
        }
        
        # CACHE THE RESULT
        set_cached_result(user_input, {
            'results': results,
            'chart_data': json.dumps(chart_data)
        })
        
        total_time = (datetime.utcnow() - start_time).total_seconds()
        logger.info(f"✅ Page rendered in {total_time:.2f} seconds (enrichment will load via AJAX)")

        return render_template("results.html", results=results, chart_data=json.dumps(chart_data))
    
    return render_template("index.html", form=form)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 🚀 UNIFIED THREAT INTELLIGENCE ENDPOINT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@main_bp.route("/scan", methods=["GET", "POST"])
@login_required
def scan_unified():
    """
    🎯 UNIFIED THREAT INTELLIGENCE ORCHESTRATOR
    
    This endpoint orchestrates ALL threat intelligence modules:
    - Automatic input detection (IP, URL, domain, hash, keyword)
    - Parallel API fetching (VirusTotal, Shodan, OTX, Google CSE)
    - ML classification (Random Forest with TF-IDF)
    - LLM analysis (Gemini/Ollama)
    - Unified results display
    
    Flow:
    1. Detect input type automatically
    2. Run all APIs in parallel
    3. Classify with ML model
    4. Generate LLM explanation
    5. Save to database
    6. Display unified results
    """
    form = InputForm()
    
    if form.validate_on_submit():
        user_input = form.input_data.data.strip()
        
        if not user_input:
            flash("Please enter a keyword, IP, URL, domain, or hash.", "warning")
            return redirect(url_for("main.scan_unified"))
        
        logger.info(f"\n{'='*80}")
        logger.info(f"🎯 UNIFIED SCAN REQUEST")
        logger.info(f"   User: {session.get('username', 'Anonymous')}")
        logger.info(f"   Input: {user_input}")
        logger.info(f"{'='*80}\n")
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # STEP 1: Check Cache
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        cached = get_cached_result(user_input)
        if cached:
            logger.info("💾 Returning cached results")
            flash("Results loaded from cache (scanned recently)", "info")
            return render_template(
                "results.html",
            # Fetch only VT, Shodan, AbuseIPDB and Google (skip OTX)        
                chart_data=cached.get('chart_data', '{}')
            )
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # STEP 2: Run Unified Orchestration Pipeline
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        try:
            orchestrated_data = orchestrate_threat_intelligence(user_input)
        except Exception as e:
            logger.error(f"❌ Orchestration failed: {e}", exc_info=True)
            flash(f"Error analyzing threat: {str(e)}", "danger")
            return redirect(url_for("main.scan_unified"))
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # STEP 3: Save to Database
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        user_ref = None
        if session.get("user_id"):
            try:
                user_ref = User.objects.get(id=session.get("user_id"))
            except Exception as e:
                logger.warning(f"User lookup failed: {e}")
        
        ioc_result = IOCResult(
            input_value=orchestrated_data['input_value'],
            type=orchestrated_data['input_type'],
            vt_report=sanitize_for_mongo(orchestrated_data.get('vt_data', {})),
            shodan_report=sanitize_for_mongo(orchestrated_data.get('shodan_data', {})),
            abuseipdb_report=sanitize_for_mongo(orchestrated_data.get('abuseipdb_data', {})),
            otx_report=sanitize_for_mongo(orchestrated_data.get('otx_data', {})),
            scraped_data=sanitize_for_mongo(orchestrated_data.get('google_data', [])),
            classification=orchestrated_data.get('classification', 'Unknown'),
            enrichment_context=sanitize_for_mongo(orchestrated_data.get('llm_analysis', {})),
            user_id=user_ref,
            timestamp=datetime.utcnow()
        )
        ioc_result.save()
        logger.info(f"💾 Saved to MongoDB with ID: {ioc_result.id}")
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # STEP 4: Format Results for Template
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        template_data = orchestrator_format_results(orchestrated_data)
        template_data['ioc_id'] = str(ioc_result.id)
        template_data['input'] = orchestrated_data['input_value']
        template_data['type'] = orchestrated_data['input_type']
        
        # Format OTX data for template compatibility
        if orchestrated_data.get('otx_data'):
            template_data['otx_formatted'] = format_otx_for_display(
                orchestrated_data['otx_data']
            )
        
        # Format Google data for template compatibility
        if orchestrated_data.get('google_data'):
            template_data['google_formatted'] = format_google_for_display(
                orchestrated_data['google_data']
            )
            template_data['scraped'] = orchestrated_data['google_data']
        
        # Chart data for dashboard
        chart_data = {
            "labels": ["Malicious", "Benign", "Informational", "Suspicious", "Unknown"],
            "values": [
                IOCResult.objects(classification="Malicious").count(),
                IOCResult.objects(classification="Benign").count(),
                IOCResult.objects(classification="Informational").count(),
                IOCResult.objects(classification="Suspicious").count(),
                IOCResult.objects(classification="Unknown").count(),
            ]
        }
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # STEP 5: Cache Results
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        set_cached_result(user_input, {
            'results': template_data,
            'chart_data': json.dumps(chart_data)
        })
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # STEP 6: Render Results
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        logger.info(f"\n{'='*80}")
        logger.info(f"✅ UNIFIED ANALYSIS COMPLETE")
        logger.info(f"   Input Type: {orchestrated_data.get('input_type')}")
        logger.info(f"   Classification: {template_data.get('classification')}")
        logger.info(f"   Total Time: {orchestrated_data.get('timing', {}).get('pipeline_total', 'N/A')}")
        logger.info(f"   Errors: {len(orchestrated_data.get('errors', []))}")
        if orchestrated_data.get('classification_details'):
            logger.info(f"   ML Confidence: {orchestrated_data['classification_details'].get('model_confidence', 'N/A')}")
        logger.info(f"{'='*80}\n")
        
        if orchestrated_data.get('errors'):
            flash(f"Note: {len(orchestrated_data['errors'])} API(s) had issues. Check logs for details.", "warning")
        
        flash(f"Analysis complete! Processed in {orchestrated_data.get('timing', {}).get('pipeline_total', 'N/A')}", "success")
        
        return render_template(
            "results.html",
            results=template_data,
            chart_data=json.dumps(chart_data)
        )
    
    # GET request - show form
    return render_template("index.html", form=form)



# ✅ NEW: AJAX Endpoint for Lazy Loading OTX Data
@main_bp.route("/api/otx/<ioc_id>", methods=["GET"])
@login_required
def get_otx_data(ioc_id):
    """
    AJAX endpoint to load OTX data after page renders
    """
    try:
        logger.info(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        logger.info(f"🔄 AJAX: /api/otx/{ioc_id} CALLED")
        logger.info(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        ioc_result = IOCResult.objects.get(id=ioc_id)
        
        # Check if OTX data already exists and is valid
        otx_data = ioc_result.otx_report or {}
        
        # If already loaded and not timed out, return cached
        if otx_data and 'timeout' not in otx_data and 'error' not in otx_data and otx_data.get('source_count', 0) > 0:
            logger.info(f"✅ Returning cached OTX data for {ioc_id}")
            return jsonify({
                'status': 'complete',
                'otx_data': otx_data,
                'otx_formatted': format_otx_for_display(otx_data)
            })
        
        # Generate OTX data now
        logger.info(f"🔄 Fetching OTX data for {ioc_id}")
        
        # Fetch with extended timeout for keywords
        from app.otx_api import otx_lookup
        
        otx_timeout = 40 if ioc_result.type == "keyword" else 20
        
        def fetch_otx():
            return otx_lookup(ioc_result.input_value, ioc_result.type)
        
        # Use ThreadPoolExecutor with timeout
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
        
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(fetch_otx)
            try:
                otx_data = future.result(timeout=otx_timeout)
            except FutureTimeoutError:
                logger.warning(f"⏱️ OTX timeout after {otx_timeout}s for {ioc_id} ({ioc_result.type})")
                otx_data = {
                    'error': 'OTX request timeout',
                    'timeout': True,
                    'threat_score': 0,
                    'classification': 'Unknown'
                }
        
        # Save to database
        ioc_result.otx_report = sanitize_for_mongo(otx_data)
        ioc_result.save()
        
        logger.info(f"✅ OTX data fetched for {ioc_id}")
        
        sanitized_otx = sanitize_for_mongo(otx_data)

        return jsonify({
            'status': 'complete',
            'otx_data': sanitized_otx,
            'otx_formatted': format_otx_for_display(sanitized_otx)
        })
    
    except Exception as e:
        logger.error(f"❌ OTX fetch error: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500


# ✅ NEW: AJAX Endpoint for ML Classification (after OTX loads)
@main_bp.route("/api/classify/<ioc_id>", methods=["GET"])
@login_required
def get_classification(ioc_id):
    """
    AJAX endpoint to run ML classification after OTX data is available
    """
    try:
        ioc_result = IOCResult.objects.get(id=ioc_id)
        
        # Check if classification already exists
        if ioc_result.classification and ioc_result.classification not in [None, 'Unknown', 'Loading...', 'Pending']:
            logger.info(f"✅ Returning cached classification for {ioc_id}: {ioc_result.classification}")
            return jsonify({
                'status': 'complete',
                'classification': ioc_result.classification
            })
        
        # Run classification now
        logger.info(f"🤖 Running ML classification for {ioc_id}")
        
        # Get all data
        vt_data = ioc_result.vt_report or {}
        shodan_data = ioc_result.shodan_report or {}
        otx_data = ioc_result.otx_report or {}
        abuseipdb_data = ioc_result.abuseipdb_report or {}
        
        # For keywords, prepare Google data
        google_formatted = None
        if ioc_result.type == "keyword" and ioc_result.scraped_data:
            google_formatted = format_google_for_display(ioc_result.scraped_data)
        
        # Run classification with detailed logging
        if classify_threat_with_details:
            classification, details = classify_threat_with_details(
                vt_data=vt_data,
                shodan_data=shodan_data,
                otx_data=otx_data,
                abuseipdb_data=abuseipdb_data,
                ioc_type=ioc_result.type,
                user_input=ioc_result.input_value,
                google_data=google_formatted
            )
            
            # Log comprehensive details
            logger.info(f"\n{'='*80}")
            logger.info(f"📊 DETAILED CLASSIFICATION RESULTS")
            logger.info(f"{'='*80}")
            logger.info(f"Input: {ioc_result.input_value} ({ioc_result.type})")
            logger.info(f"Final Classification: {classification}")
            logger.info(f"\n🤖 Model Metrics:")
            logger.info(f"   Confidence: {details.get('model_confidence', 0):.2%}")
            logger.info(f"   Threshold Met (0.75): {details.get('confidence_threshold_met', False)}")
            
            if 'api_scores' in details:
                api_scores = details['api_scores']
                logger.info(f"\n📡 API Scores:")
                logger.info(f"   VirusTotal: {api_scores.get('vt_score', 0):.1f}/100")
                logger.info(f"   OTX: {api_scores.get('otx_score', 0):.1f}/100")
                logger.info(f"   Shodan: {api_scores.get('shodan_score', 0):.1f}/100")
                logger.info(f"   AbuseIPDB: {api_scores.get('abuseipdb_score', 0):.1f}/100")
                logger.info(f"   Weighted Total: {api_scores.get('weighted_total', 0):.1f}/100")
                logger.info(f"   Weights: VT={api_scores.get('weights_used', {}).get('vt', 0):.0%}, "
                          f"OTX={api_scores.get('weights_used', {}).get('otx', 0):.0%}, "
                          f"Shodan={api_scores.get('weights_used', {}).get('shodan', 0):.0%}, "
                          f"AbuseIPDB={api_scores.get('weights_used', {}).get('abuseipdb', 0):.0%}")
            
            if 'context_analysis' in details:
                context = details['context_analysis']
                logger.info(f"\n🧠 Context Analysis:")
                logger.info(f"   Benign patterns found: {context.get('benign_score', 0)}")
                logger.info(f"   Malicious patterns found: {context.get('malicious_score', 0)}")
                logger.info(f"   Is benign context: {context.get('is_benign_context', False)}")
                logger.info(f"   Is malicious context: {context.get('is_malicious_context', False)}")
            
            logger.info(f"\n📋 Reasoning:")
            for reason in details.get('reasoning', []):
                logger.info(f"   - {reason}")
            logger.info(f"{'='*80}\n")
            
            # Store detailed results in enrichment context
            if not ioc_result.enrichment_context:
                ioc_result.enrichment_context = {}
            ioc_result.enrichment_context['classification_details'] = details
        else:
            # Fallback to simple classification
            classification = classify_threat(
                vt_data=vt_data,
                shodan_data=shodan_data,
                otx_data=otx_data,
                abuseipdb_data=abuseipdb_data,
                ioc_type=ioc_result.type,
                user_input=ioc_result.input_value,
                google_data=google_formatted
            )
        
        # Save to database
        logger.info(f"💾 Saving classification to database...")
        ioc_result.classification = classification
        ioc_result.save()
        logger.info(f"✅ Classification SAVED to database: {classification}")

    # ✅ Verify it was saved
        ioc_result.reload()
        logger.info(f"✅ Verified in DB: {ioc_result.classification}")
        
        return jsonify({
            'status': 'complete',
            'classification': classification
        })
    
    except Exception as e:
        logger.error(f"❌ Classification error: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500
# ✅ NEW: AJAX Endpoint for Lazy Loading Enrichment
@main_bp.route("/api/enrichment/<ioc_id>", methods=["GET"])
@login_required
def get_enrichment(ioc_id):
    """AJAX endpoint to load AI enrichment after page renders"""
    try:
        logger.info(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        logger.info(f"🧠 AJAX: /api/enrichment/{ioc_id} CALLED")
        logger.info(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        
        ioc_result = IOCResult.objects.get(id=ioc_id)
        
        logger.info(f"📊 Current IOC state:")
        logger.info(f"   Classification: {ioc_result.classification}")
        logger.info(f"   Has OTX data: {bool(ioc_result.otx_report)}")
        logger.info(f"   Has enrichment: {bool(ioc_result.enrichment_context)}")
        
        # ✅ CRITICAL: Check if classification is valid
        if not ioc_result.classification or ioc_result.classification in ['Loading...', None, 'Unknown', 'Pending', '']:
            current_cls = ioc_result.classification or 'None'
            logger.warning(f"⚠️ ⚠️ ⚠️ Classification NOT ready: '{current_cls}' - Returning 400")
            return jsonify({
                'status': 'waiting',
                'error': f'Classification is {current_cls}. Waiting for ML analysis to complete.',
                'current_classification': current_cls
            }), 400
        
        logger.info(f"✅ Classification is VALID: {ioc_result.classification}")
        
        # Check if enrichment already exists
        enrichment = ioc_result.enrichment_context or {}
        
        # ✅ If enrichment exists but contains invalid text, regenerate
        if enrichment and not enrichment.get('loading'):
            summary = enrichment.get('summary', '')
            why_malicious = enrichment.get('why_malicious', '')
            
            # Check for invalid classification strings in text
            invalid_strings = ['Pending', 'Loading...', 'Loading']
            has_invalid = any(inv in summary or inv in why_malicious for inv in invalid_strings)
            
            # Also check if enrichment is essentially empty
            is_empty = not summary and not why_malicious and not enrichment.get('recommendation')
            
            if has_invalid:
                logger.warning(f"⚠️ Enrichment contains invalid text (Pending/Loading), regenerating...")
                enrichment = {}  # Force regeneration
            elif is_empty:
                logger.warning(f"⚠️ Enrichment is empty, regenerating...")
                enrichment = {}  # Force regeneration
            else:
                logger.info(f"✅ Returning cached enrichment for {ioc_id}")
                logger.info(f"📦 Enrichment data keys: {list(enrichment.keys())}")
                logger.info(f"   Summary: {enrichment.get('summary', 'N/A')[:100]}...")
                logger.info(f"   Why Malicious: {enrichment.get('why_malicious', 'N/A')[:100]}...")
                logger.info(f"   Recommendation: {enrichment.get('recommendation', 'N/A')[:100]}...")
                return jsonify({
                    'status': 'complete',
                    'enrichment': enrichment
                })
        
        # Generate enrichment now
        logger.info(f"🔍 Generating fresh enrichment for {ioc_id}")
        
        ENABLE_AI_ENRICHMENT = os.getenv('ENABLE_ENRICHMENT', 'true').lower() == 'true'
        
        if ENABLE_AI_ENRICHMENT:
            try:
                from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
                
                def generate_enrichment():
                    return enrich_threat_intelligence(
                        ioc_value=ioc_result.input_value,
                        ioc_type=ioc_result.type,
                        vt_data=ioc_result.vt_report,
                        shodan_data=ioc_result.shodan_report,
                        otx_data=ioc_result.otx_report,
                        classification=ioc_result.classification
                    )
                
                # 30 second timeout for AI enrichment
                with ThreadPoolExecutor(max_workers=1) as executor:
                    future = executor.submit(generate_enrichment)
                    try:
                        enrichment = future.result(timeout=30)
                        logger.info(f"✅ Enrichment generated successfully")
                    except FutureTimeoutError:
                        logger.warning(f"⏱️ Enrichment timeout after 30s")
                        enrichment = {
                            'summary': f'AI analysis timed out. This {ioc_result.type} was classified as {ioc_result.classification}.',
                            'why_malicious': f'This {ioc_result.type} was classified as {ioc_result.classification} based on threat intelligence data. AI-powered detailed analysis was unavailable due to timeout.',
                            'key_indicators': ['Classified using ML model', 'Multiple threat intelligence sources consulted'],
                            'recommendation': 'Review threat intelligence data manually for detailed analysis.',
                            'confidence': 'Medium'
                        }
                        
            except Exception as e:
                logger.error(f"❌ Enrichment generation error: {e}", exc_info=True)
                enrichment = {
                    'summary': f'This {ioc_result.type} was classified as {ioc_result.classification}.',
                    'why_malicious': f'Classification: {ioc_result.classification}. AI analysis encountered an error.',
                    'key_indicators': ['Manual review recommended'],
                    'recommendation': 'Refer to threat intelligence reports above.',
                    'confidence': 'Low',
                    'error': str(e)
                }
        else:
            enrichment = {
                'summary': f'This {ioc_result.type} was classified as {ioc_result.classification}.',
                'why_malicious': f'AI enrichment is disabled. Classification: {ioc_result.classification}.',
                'key_indicators': ['Classification based on threat intelligence'],
                'recommendation': 'Review threat intelligence data above.',
                'confidence': 'Medium'
            }
        
        # Save enrichment to database
        ioc_result.enrichment_context = enrichment
        ioc_result.save()
        
        logger.info(f"💾 Enrichment saved to database for {ioc_id}")
        
        return jsonify({
            'status': 'complete',
            'enrichment': enrichment
        })
    
    except IOCResult.DoesNotExist:
        logger.error(f"❌ IOC {ioc_id} not found")
        return jsonify({
            'status': 'error',
            'error': 'IOC not found'
        }), 404
    except Exception as e:
        logger.error(f"❌ Enrichment API error for {ioc_id}: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

# ---- FEEDBACK (Protected) ----
@main_bp.route("/feedback/<ioc_id>", methods=["POST"])
@login_required
def feedback(ioc_id):
    feedback_value = request.form.get("feedback")
    if feedback_value not in ["Malicious", "Benign", "Informational"]:
        return jsonify({"error": "Invalid feedback"}), 400

    try:
        ioc = IOCResult.objects.get(id=ioc_id)
        new_feedback = Feedback(ioc_id=ioc, correct_classification=feedback_value)
        new_feedback.save()
        return jsonify({"message": "Feedback saved"})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


# ---- EXPORT ROUTES (Protected) ----
# ---- EXPORT ROUTES (ENHANCED WITH AI ENRICHMENT) ----
@main_bp.route("/export/<format>/<ioc_id>")
@login_required
def export(format, ioc_id):
    """Export single result in multiple formats with AI enrichment"""
    try:
        result = IOCResult.objects.get(id=ioc_id)
    except Exception as e:
        return jsonify({"error": f"Result not found: {str(e)}"}), 404
    
    # ✅ Get enrichment data (trigger generation if needed)
    enrichment = result.enrichment_context
    if not enrichment or enrichment.get('loading'):
        # Try to generate it now for export
        try:
            from app.enrichment import enrich_threat_intelligence
            enrichment = enrich_threat_intelligence(
                ioc_value=result.input_value,
                ioc_type=result.type,
                vt_data=result.vt_report,
                shodan_data=result.shodan_report,
                otx_data=result.otx_report,
                classification=result.classification
            )
            result.enrichment_context = enrichment
            result.save()
        except Exception as e:
            logger.error(f"Failed to generate enrichment for export: {e}")
            enrichment = {
                'summary': 'AI analysis unavailable',
                'recommendation': 'Manual review required',
                'confidence': 'Low'
            }
    
    vt_data = result.vt_report or {}
    shodan_data = result.shodan_report or {}
    otx_data = result.otx_report or {}
    abuseipdb_data = result.abuseipdb_report or {}
    google_formatted = format_google_for_display(result.scraped_data) if result.scraped_data else None

    # JSON Export
    if format == "json":
        return jsonify({
            "metadata": {
                "id": str(result.id),
                "generated_at": datetime.utcnow().isoformat(),
                "report_version": "2.0"
            },
            "ioc_details": {
                "input": result.input_value,
                "type": result.type,
                "classification": result.classification,
                "timestamp": result.timestamp.isoformat() if result.timestamp else None
            },
            "ai_analysis": {
                "summary": enrichment.get('summary', 'N/A'),
                "detailed_explanation": enrichment.get('why_malicious', 'N/A'),
                "key_indicators": enrichment.get('key_indicators', []),
                "recommendation": enrichment.get('recommendation', 'N/A'),
                "confidence_level": enrichment.get('confidence', 'Unknown'),
                "risk_score": enrichment.get('risk_score', 0),
                "sources_analyzed": enrichment.get('sources_analyzed', [])
            },
            "threat_intelligence": {
                "virustotal": vt_data,
                "shodan": shodan_data,
                "alienvault_otx": otx_data,
                "abuseipdb": abuseipdb_data,
                "google_search": {
                    "summary": google_formatted,
                    "raw_results": result.scraped_data
                }
            },
            "scraped_data": result.scraped_data
        })
    
    # CSV Export
    elif format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["=" * 80])
        writer.writerow(["THREAT INTELLIGENCE REPORT"])
        writer.writerow(["Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
        writer.writerow(["=" * 80])
        writer.writerow([])
        
        # IOC Information
        writer.writerow(["━" * 80])
        writer.writerow(["INDICATOR OF COMPROMISE (IOC)"])
        writer.writerow(["━" * 80])
        writer.writerow(["ID", str(result.id)])
        writer.writerow(["Input", result.input_value])
        writer.writerow(["Type", result.type])
        writer.writerow(["Classification", result.classification])
        writer.writerow(["Scan Date", result.timestamp.strftime("%Y-%m-%d %H:%M:%S") if result.timestamp else "N/A"])
        writer.writerow([])
        
        # AI Analysis Section
        writer.writerow(["━" * 80])
        writer.writerow(["🤖 AI THREAT ANALYSIS"])
        writer.writerow(["━" * 80])
        writer.writerow([])
        
        writer.writerow(["Summary"])
        writer.writerow([enrichment.get('summary', 'N/A')])
        writer.writerow([])
        
        writer.writerow(["Detailed Analysis"])
        # Split long text into multiple rows
        explanation = enrichment.get('why_malicious', 'N/A')
        for line in explanation.split('\n'):
            if line.strip():
                writer.writerow([line.strip()])
        writer.writerow([])
        
        writer.writerow(["Key Threat Indicators"])
        indicators = enrichment.get('key_indicators', [])
        if indicators:
            for idx, indicator in enumerate(indicators, 1):
                writer.writerow([f"  {idx}.", indicator])
        else:
            writer.writerow(["  No specific indicators identified"])
        writer.writerow([])
        
        writer.writerow(["Recommended Action"])
        writer.writerow([enrichment.get('recommendation', 'N/A')])
        writer.writerow([])
        
        writer.writerow(["Confidence Level", enrichment.get('confidence', 'Unknown')])
        writer.writerow(["Risk Score", f"{enrichment.get('risk_score', 0)}/100"])
        writer.writerow(["Sources Analyzed", ", ".join(enrichment.get('sources_analyzed', []))])
        writer.writerow([])
        
        # VirusTotal Report
        writer.writerow(["━" * 80])
        writer.writerow(["VIRUSTOTAL ANALYSIS"])
        writer.writerow(["━" * 80])
        if vt_data and 'error' not in vt_data:
            stats = vt_data.get("last_analysis_stats", {})
            if not stats:
                stats = vt_data.get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            
            if stats:
                writer.writerow(["Malicious Detections", stats.get("malicious", 0)])
                writer.writerow(["Suspicious Detections", stats.get("suspicious", 0)])
                writer.writerow(["Undetected", stats.get("undetected", 0)])
                writer.writerow(["Harmless", stats.get("harmless", 0)])
            else:
                writer.writerow(["No detection statistics available"])
        else:
            writer.writerow(["No VirusTotal data available"])
        writer.writerow([])
        
        # Shodan Report
        writer.writerow(["━" * 80])
        writer.writerow(["SHODAN NETWORK INTELLIGENCE"])
        writer.writerow(["━" * 80])
        if shodan_data and 'error' not in shodan_data:
            writer.writerow(["IP Address", shodan_data.get("ip_str", shodan_data.get("ip", "N/A"))])
            writer.writerow(["Organization", shodan_data.get("org", "N/A")])
            writer.writerow(["ISP", shodan_data.get("isp", "N/A")])
            writer.writerow(["Country", shodan_data.get("country_name", "N/A")])
            
            ports = shodan_data.get("ports", [])
            if not ports and "data" in shodan_data:
                ports = [item.get("port") for item in shodan_data["data"] if "port" in item]
            
            writer.writerow(["Open Ports", ", ".join(map(str, ports)) if ports else "None detected"])
        else:
            writer.writerow(["No Shodan data available"])
        writer.writerow([])

        # AbuseIPDB Report
        writer.writerow(["━" * 80])
        writer.writerow(["ABUSEIPDB REPUTATION"])
        writer.writerow(["━" * 80])
        if abuseipdb_data and 'error' not in abuseipdb_data:
            writer.writerow(["Threat Score", f"{abuseipdb_data.get('threat_score', 0)}/100"])
            writer.writerow(["Abuse Confidence", f"{abuseipdb_data.get('abuse_confidence_score', 0)}%"])
            writer.writerow(["Classification", abuseipdb_data.get('classification', 'Unknown')])
            writer.writerow(["Total Reports", abuseipdb_data.get('total_reports', 0)])
            writer.writerow(["Last Reported", abuseipdb_data.get('last_reported', 'N/A')])
            writer.writerow(["Country", abuseipdb_data.get('country_code', 'Unknown')])
            writer.writerow(["Usage Type", abuseipdb_data.get('usage_type', 'Unknown')])
            writer.writerow(["ISP", abuseipdb_data.get('isp', 'Unknown')])
            categories = abuseipdb_data.get('categories') or abuseipdb_data.get('details', {}).get('categories', {})
            categories = categories if isinstance(categories, dict) else {}
            if categories:
                writer.writerow([])
                writer.writerow(["Category Breakdown"])
                for name, count in categories.items():
                    writer.writerow([f"  • {name}", f"{count} reports"])
        else:
            writer.writerow(["No AbuseIPDB data available or lookup not applicable"])
        writer.writerow([])

        # Google OSINT
        writer.writerow(["━" * 80])
        writer.writerow(["GOOGLE SEARCH INTELLIGENCE"])
        writer.writerow(["━" * 80])
        if google_formatted:
            writer.writerow(["Search Results Found", google_formatted.get('result_count', 0)])
            writer.writerow(["Threat Score", f"{google_formatted.get('threat_score', 0)}/100"])
            writer.writerow(["Risk Classification", google_formatted.get('classification', 'N/A')])
            writer.writerow(["Unique Domains", google_formatted.get('total_domains', 0)])

            threat_indicators = google_formatted.get('threat_indicators', [])
            if threat_indicators:
                writer.writerow([])
                writer.writerow(["Threat Indicators"])
                for indicator in threat_indicators:
                    writer.writerow([f"  • {indicator}"])

            safe_indicators = google_formatted.get('safe_indicators', [])
            if safe_indicators:
                writer.writerow([])
                writer.writerow(["Safe Indicators"])
                for indicator in safe_indicators:
                    writer.writerow([f"  • {indicator}"])

            domains = google_formatted.get('unique_domains', [])
            if domains:
                writer.writerow([])
                writer.writerow(["Notable Domains"])
                for domain in domains[:10]:
                    writer.writerow([f"  • {domain}"])

            top_results = google_formatted.get('top_results', [])
            if top_results:
                writer.writerow([])
                writer.writerow(["Top Search Results"])
                for entry in top_results:
                    writer.writerow([entry.get('title', 'No title')])
                    writer.writerow([entry.get('url', '#')])
                    writer.writerow([entry.get('snippet', 'No description')])
                    writer.writerow([])
        else:
            writer.writerow(["No Google search data available"])
        writer.writerow([])
        
        # OTX Report
        writer.writerow(["━" * 80])
        writer.writerow(["ALIENVAULT OTX THREAT INTELLIGENCE"])
        writer.writerow(["━" * 80])
        if otx_data and 'error' not in otx_data:
            writer.writerow(["Threat Score", f"{otx_data.get('threat_score', 0)}/100"])
            writer.writerow(["Classification", otx_data.get('classification', 'Unknown')])
            writer.writerow(["Threat Pulses", otx_data.get('source_count', 0)])
            
            details = otx_data.get('details', {})
            if details:
                writer.writerow(["Risk Level", details.get('risk_level', 'Unknown')])
                
                malware = details.get('malware_families', [])
                if malware:
                    writer.writerow(["Malware Families", ", ".join(malware[:5])])
                
                tags = details.get('top_tags', [])
                if tags:
                    writer.writerow(["Top Tags", ", ".join(tags[:10])])
        else:
            writer.writerow(["No OTX data available"])
        
        writer.writerow([])
        writer.writerow(["=" * 80])
        writer.writerow(["END OF REPORT"])
        writer.writerow(["=" * 80])
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"threat_report_{result.input_value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        

        
    
    # Excel Export
    elif format == "excel":
        wb = openpyxl.Workbook()
        
        # ============================================
        # WORKSHEET 1: EXECUTIVE SUMMARY
        # ============================================
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        # Styling
        header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=14)
        title_font = Font(bold=True, size=20, color="4361ee")
        section_font = Font(bold=True, size=12, color="4cc9f0")
        
        # Title
        ws_summary['A1'] = "🛡️ THREAT INTELLIGENCE REPORT"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws_summary.merge_cells('A2:D2')
        
        row = 4
        
        # IOC Details
        ws_summary[f'A{row}'] = "INDICATOR OF COMPROMISE"
        ws_summary[f'A{row}'].font = section_font
        ws_summary.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ioc_data = [
            ["Input Value", result.input_value],
            ["Type", result.type],
            ["Classification", result.classification],
            ["Risk Score", f"{enrichment.get('risk_score', 0)}/100"],
            ["Scan Date", result.timestamp.strftime("%Y-%m-%d %H:%M:%S") if result.timestamp else "N/A"]
        ]

        if abuseipdb_data and 'error' not in abuseipdb_data:
            ioc_data.append(["AbuseIPDB Confidence", f"{abuseipdb_data.get('abuse_confidence_score', 0)}%"])
            ioc_data.append(["AbuseIPDB Reports", abuseipdb_data.get('total_reports', 0)])

        if google_formatted:
            ioc_data.append(["Google Threat Score", f"{google_formatted.get('threat_score', 0)}/100"])
            ioc_data.append(["Google Classification", google_formatted.get('classification', 'Unknown')])
        
        for item in ioc_data:
            ws_summary[f'A{row}'] = item[0]
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = item[1]
            ws_summary.merge_cells(f'B{row}:D{row}')
            row += 1
        
        row += 1
        
        # AI Analysis
        ws_summary[f'A{row}'] = "🤖 AI THREAT ANALYSIS"
        ws_summary[f'A{row}'].font = section_font
        ws_summary.merge_cells(f'A{row}:D{row}')
        row += 1
        
        ws_summary[f'A{row}'] = "Summary"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        ws_summary[f'A{row}'] = enrichment.get('summary', 'N/A')
        ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws_summary.merge_cells(f'A{row}:D{row}')
        ws_summary.row_dimensions[row].height = 40
        row += 2
        
        ws_summary[f'A{row}'] = "Key Indicators"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        
        indicators = enrichment.get('key_indicators', [])
        if indicators:
            for indicator in indicators:
                ws_summary[f'A{row}'] = f"• {indicator}"
                ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws_summary.merge_cells(f'A{row}:D{row}')
                row += 1
        else:
            ws_summary[f'A{row}'] = "No specific indicators"
            row += 1
        
        row += 1
        ws_summary[f'A{row}'] = "Recommended Action"
        ws_summary[f'A{row}'].font = Font(bold=True, color="e74c3c" if result.classification == "Malicious" else "2ecc71")
        row += 1
        ws_summary[f'A{row}'] = enrichment.get('recommendation', 'N/A')
        ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws_summary.merge_cells(f'A{row}:D{row}')
        ws_summary.row_dimensions[row].height = 60
        
        # Set column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 50
        ws_summary.column_dimensions['C'].width = 20
        ws_summary.column_dimensions['D'].width = 20
        
        # ============================================
        # WORKSHEET 2: DETAILED ANALYSIS
        # ============================================
        ws_detail = wb.create_sheet("Detailed Analysis")
        
        row = 1
        ws_detail[f'A{row}'] = "DETAILED THREAT ANALYSIS"
        ws_detail[f'A{row}'].font = title_font
        ws_detail.merge_cells(f'A{row}:B{row}')
        row += 2
        
        ws_detail[f'A{row}'] = "Why This Was Classified as " + result.classification
        ws_detail[f'A{row}'].font = section_font
        ws_detail.merge_cells(f'A{row}:B{row}')
        row += 1
        
        explanation = enrichment.get('why_malicious', 'No detailed explanation available')
        ws_detail[f'A{row}'] = explanation
        ws_detail[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_detail.merge_cells(f'A{row}:B{row}')
        ws_detail.row_dimensions[row].height = 150
        row += 2
        
        ws_detail[f'A{row}'] = "Confidence Level"
        ws_detail[f'A{row}'].font = Font(bold=True)
        ws_detail[f'B{row}'] = enrichment.get('confidence', 'Unknown')
        row += 1
        
        ws_detail[f'A{row}'] = "Sources Analyzed"
        ws_detail[f'A{row}'].font = Font(bold=True)
        ws_detail[f'B{row}'] = ", ".join(enrichment.get('sources_analyzed', []))
        
        ws_detail.column_dimensions['A'].width = 25
        ws_detail.column_dimensions['B'].width = 80
        
        # ============================================
        # WORKSHEET 3: VIRUSTOTAL
        # ============================================
        ws_vt = wb.create_sheet("VirusTotal")
        
        row = 1
        ws_vt[f'A{row}'] = "VIRUSTOTAL ANALYSIS"
        ws_vt[f'A{row}'].font = title_font
        row += 2
        
        if vt_data and 'error' not in vt_data:
            stats = vt_data.get("last_analysis_stats", {})
            if not stats:
                stats = vt_data.get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            
            if stats:
                ws_vt[f'A{row}'] = "Detection Category"
                ws_vt[f'B{row}'] = "Count"
                ws_vt[f'A{row}'].font = header_font
                ws_vt[f'B{row}'].font = header_font
                ws_vt[f'A{row}'].fill = header_fill
                ws_vt[f'B{row}'].fill = header_fill
                row += 1
                
                for key, value in stats.items():
                    ws_vt[f'A{row}'] = key.capitalize()
                    ws_vt[f'B{row}'] = value
                    row += 1
        else:
            ws_vt[f'A{row}'] = "No VirusTotal data available"
        
        ws_vt.column_dimensions['A'].width = 30
        ws_vt.column_dimensions['B'].width = 20
        
        # ============================================
        # WORKSHEET 4: SHODAN
        # ============================================
        ws_shodan = wb.create_sheet("Shodan")
        
        row = 1
        ws_shodan[f'A{row}'] = "SHODAN NETWORK INTELLIGENCE"
        ws_shodan[f'A{row}'].font = title_font
        row += 2
        
        if shodan_data and 'error' not in shodan_data:
            shodan_rows = [
                ["IP Address", shodan_data.get("ip_str", shodan_data.get("ip", "N/A"))],
                ["Organization", shodan_data.get("org", "N/A")],
                ["ISP", shodan_data.get("isp", "N/A")],
                ["Country", shodan_data.get("country_name", "N/A")],
                ["City", shodan_data.get("city", "N/A")]
            ]
            
            for item in shodan_rows:
                ws_shodan[f'A{row}'] = item[0]
                ws_shodan[f'A{row}'].font = Font(bold=True)
                ws_shodan[f'B{row}'] = item[1]
                row += 1
            
            row += 1
            shodan_ports = shodan_data.get("ports", [])
            if not shodan_ports and "data" in shodan_data:
                shodan_ports = [item.get("port") for item in shodan_data["data"] if "port" in item]
            
            ws_shodan[f'A{row}'] = "Open Ports"
            ws_shodan[f'A{row}'].font = Font(bold=True)
            ws_shodan[f'B{row}'] = ", ".join(map(str, shodan_ports)) if shodan_ports else "None"
        else:
            ws_shodan[f'A{row}'] = "No Shodan data available"
        
        ws_shodan.column_dimensions['A'].width = 30
        ws_shodan.column_dimensions['B'].width = 50
        
        # ============================================
        # WORKSHEET 5: OTX
        # ============================================
        ws_otx = wb.create_sheet("AlienVault OTX")
        
        row = 1
        ws_otx[f'A{row}'] = "ALIENVAULT OTX THREAT INTELLIGENCE"
        ws_otx[f'A{row}'].font = title_font
        row += 2
        
        if otx_data and 'error' not in otx_data:
            otx_rows = [
                ["Threat Score", f"{otx_data.get('threat_score', 0)}/100"],
                ["Classification", otx_data.get('classification', 'Unknown')],
                ["Threat Pulses", otx_data.get('source_count', 0)]
            ]
            
            details = otx_data.get('details', {})
            if details:
                otx_rows.append(["Risk Level", details.get('risk_level', 'Unknown')])
            
            for item in otx_rows:
                ws_otx[f'A{row}'] = item[0]
                ws_otx[f'A{row}'].font = Font(bold=True)
                ws_otx[f'B{row}'] = item[1]
                row += 1
            
            if details:
                row += 1
                malware = details.get('malware_families', [])
                if malware:
                    ws_otx[f'A{row}'] = "Malware Families"
                    ws_otx[f'A{row}'].font = Font(bold=True)
                    row += 1
                    for family in malware[:10]:
                        ws_otx[f'A{row}'] = f"• {family}"
                        row += 1
                
                row += 1
                tags = details.get('top_tags', [])
                if tags:
                    ws_otx[f'A{row}'] = "Top Tags"
                    ws_otx[f'A{row}'].font = Font(bold=True)
                    ws_otx[f'B{row}'] = ", ".join(tags[:15])
        else:
            ws_otx[f'A{row}'] = "No OTX data available"
        
        ws_otx.column_dimensions['A'].width = 30
        ws_otx.column_dimensions['B'].width = 60

        # ============================================
        # WORKSHEET 6: ABUSEIPDB
        # ============================================
        ws_abuse = wb.create_sheet("AbuseIPDB")
        row = 1
        ws_abuse[f'A{row}'] = "ABUSEIPDB REPUTATION"
        ws_abuse[f'A{row}'].font = title_font
        row += 2

        if abuseipdb_data and 'error' not in abuseipdb_data:
            abuse_rows = [
                ["Threat Score", f"{abuseipdb_data.get('threat_score', 0)}/100"],
                ["Abuse Confidence", f"{abuseipdb_data.get('abuse_confidence_score', 0)}%"],
                ["Classification", abuseipdb_data.get('classification', 'Unknown')],
                ["Total Reports", abuseipdb_data.get('total_reports', 0)],
                ["Distinct Reporters", abuseipdb_data.get('details', {}).get('distinct_reporters', abuseipdb_data.get('total_reports', 0))],
                ["Last Reported", abuseipdb_data.get('last_reported', 'N/A')],
                ["Country", abuseipdb_data.get('country_code', 'Unknown')],
                ["Usage Type", abuseipdb_data.get('usage_type', 'Unknown')],
                ["ISP", abuseipdb_data.get('isp', 'Unknown')],
                ["Domain", abuseipdb_data.get('domain', 'Unknown')]
            ]

            for label, value in abuse_rows:
                ws_abuse[f'A{row}'] = label
                ws_abuse[f'A{row}'].font = Font(bold=True)
                ws_abuse[f'B{row}'] = value
                row += 1

            categories = abuseipdb_data.get('categories') or abuseipdb_data.get('details', {}).get('categories', {})
            categories = categories if isinstance(categories, dict) else {}
            if categories:
                row += 1
                ws_abuse[f'A{row}'] = "Category Breakdown"
                ws_abuse[f'A{row}'].font = section_font
                ws_abuse.merge_cells(f'A{row}:B{row}')
                row += 1

                for name, count in categories.items():
                    ws_abuse[f'A{row}'] = f"• {name}"
                    ws_abuse[f'B{row}'] = f"{count} reports"
                    row += 1
        else:
            ws_abuse[f'A{row}'] = "No AbuseIPDB data available or lookup not applicable"

        ws_abuse.column_dimensions['A'].width = 35
        ws_abuse.column_dimensions['B'].width = 45

        # ============================================
        # WORKSHEET 7: GOOGLE OSINT
        # ============================================
        ws_google = wb.create_sheet("Google OSINT")
        row = 1
        ws_google[f'A{row}'] = "GOOGLE SEARCH INTELLIGENCE"
        ws_google[f'A{row}'].font = title_font
        row += 2

        if google_formatted:
            google_rows = [
                ["Search Results", google_formatted.get('result_count', 0)],
                ["Threat Score", f"{google_formatted.get('threat_score', 0)}/100"],
                ["Classification", google_formatted.get('classification', 'Unknown')],
                ["Risk Level", google_formatted.get('risk_level', 'Unknown')],
                ["Unique Domains", google_formatted.get('total_domains', 0)]
            ]

            for label, value in google_rows:
                ws_google[f'A{row}'] = label
                ws_google[f'A{row}'].font = Font(bold=True)
                ws_google[f'B{row}'] = value
                row += 1

            threat_indicators = google_formatted.get('threat_indicators', [])
            if threat_indicators:
                row += 1
                ws_google[f'A{row}'] = "Threat Indicators"
                ws_google[f'A{row}'].font = section_font
                ws_google.merge_cells(f'A{row}:B{row}')
                row += 1
                for indicator in threat_indicators:
                    ws_google[f'A{row}'] = f"• {indicator}"
                    ws_google.merge_cells(f'A{row}:B{row}')
                    row += 1

            safe_indicators = google_formatted.get('safe_indicators', [])
            if safe_indicators:
                row += 1
                ws_google[f'A{row}'] = "Safe Indicators"
                ws_google[f'A{row}'].font = section_font
                ws_google.merge_cells(f'A{row}:B{row}')
                row += 1
                for indicator in safe_indicators:
                    ws_google[f'A{row}'] = f"• {indicator}"
                    ws_google.merge_cells(f'A{row}:B{row}')
                    row += 1

            top_results = google_formatted.get('top_results', [])
            if top_results:
                row += 1
                ws_google[f'A{row}'] = "Top Search Results"
                ws_google[f'A{row}'].font = section_font
                ws_google.merge_cells(f'A{row}:B{row}')
                row += 1

                ws_google[f'A{row}'] = "Title"
                ws_google[f'B{row}'] = "URL / Snippet"
                ws_google[f'A{row}'].font = header_font
                ws_google[f'B{row}'].font = header_font
                ws_google[f'A{row}'].fill = header_fill
                ws_google[f'B{row}'].fill = header_fill
                row += 1

                for entry in top_results:
                    ws_google[f'A{row}'] = entry.get('title', 'No title')
                    ws_google[f'B{row}'] = entry.get('url', '#')
                    row += 1
                    ws_google[f'A{row}'] = "Snippet"
                    ws_google[f'A{row}'].font = Font(italic=True)
                    ws_google[f'B{row}'] = entry.get('snippet', 'No description')
                    ws_google[f'B{row}'].alignment = Alignment(wrap_text=True)
                    row += 1
        else:
            ws_google[f'A{row}'] = "No Google search intelligence available"

        ws_google.column_dimensions['A'].width = 35
        ws_google.column_dimensions['B'].width = 65
        
        # Save Excel file
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"threat_report_{result.input_value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    
    # PDF Export
    elif format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        # Custom Styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=28,
            textColor=colors.HexColor('#4361ee'),
            spaceAfter=10,
            alignment=1,
            fontName='Helvetica-Bold'
        )
        
        section_style = ParagraphStyle(
            'SectionHeader',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#4cc9f0'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.gray,
            alignment=1,
            spaceAfter=20
        )
        
        # Title
        story.append(Paragraph("🛡️ THREAT INTELLIGENCE REPORT", title_style))
        story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}", subtitle_style))
        story.append(Spacer(1, 0.2*inch))
        
        # IOC Summary Table
        story.append(Paragraph("INDICATOR OF COMPROMISE", section_style))
        
        ioc_data = [
            ['Field', 'Value'],
            ['Input', result.input_value],
            ['Type', result.type],
            ['Classification', result.classification],
            ['Risk Score', f"{enrichment.get('risk_score', 0)}/100"],
            ['Scan Date', result.timestamp.strftime("%Y-%m-%d %H:%M:%S") if result.timestamp else "N/A"]
        ]
        
        ioc_table = Table(ioc_data, colWidths=[2*inch, 4.5*inch])
        ioc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4361ee')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(ioc_table)
        story.append(Spacer(1, 0.3*inch))
        
        # AI Analysis Section
        story.append(Paragraph("🤖 AI THREAT ANALYSIS", section_style))
        
        # Summary
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=11,
            leading=16,
            spaceAfter=12,
            leftIndent=20,
            rightIndent=20,
            backColor=colors.HexColor('#f0f8ff')
        )
        story.append(Paragraph(f"<b>Summary:</b> {enrichment.get('summary', 'N/A')}", summary_style))
        story.append(Spacer(1, 0.1*inch))
        
        # Key Indicators
        story.append(Paragraph("<b>Key Threat Indicators:</b>", styles['Normal']))
        story.append(Spacer(1, 0.05*inch))
        
        indicators = enrichment.get('key_indicators', [])
        if indicators:
            indicator_data = [['#', 'Indicator']]
            for idx, indicator in enumerate(indicators, 1):
                indicator_data.append([str(idx), indicator])
            
            indicator_table = Table(indicator_data, colWidths=[0.5*inch, 6*inch])
            indicator_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4cc9f0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(indicator_table)
        else:
            story.append(Paragraph("No specific indicators identified", styles['Normal']))
        
        story.append(Spacer(1, 0.2*inch))
        
        # Recommendation Box
        recommendation = enrichment.get('recommendation', 'N/A')
        
        # Color code based on classification
        if result.classification == "Malicious":
            rec_color = colors.HexColor('#ffe6e6')
            border_color = colors.HexColor('#e74c3c')
        elif result.classification == "Benign":
            rec_color = colors.HexColor('#e6f7e6')
            border_color = colors.HexColor('#2ecc71')
        else:
            rec_color = colors.HexColor('#fff8e6')
            border_color = colors.HexColor('#f39c12')
        
        rec_style = ParagraphStyle(
            'Recommendation',
            parent=styles['Normal'],
            fontSize=11,
            leading=16,
            backColor=rec_color
        )
        
        rec_data = [
            [Paragraph("<b>🛡️ RECOMMENDED ACTION</b>", styles['Heading3'])],
            [Paragraph(recommendation, rec_style)]
        ]
        
        rec_table = Table(rec_data, colWidths=[6.5*inch])
        rec_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), border_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            ('BOX', (0, 0), (-1, -1), 2, border_color),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(rec_table)
        story.append(Spacer(1, 0.1*inch))
        
        # Confidence and Risk Score
        meta_data = [
            ['Confidence Level', enrichment.get('confidence', 'Unknown')],
            ['Risk Score', f"{enrichment.get('risk_score', 0)}/100"],
            ['Sources Analyzed', ", ".join(enrichment.get('sources_analyzed', []))]
        ]
        
        meta_table = Table(meta_data, colWidths=[2*inch, 4.5*inch])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Detailed Explanation (if long, add page break before)
        explanation = enrichment.get('why_malicious', '')
        if explanation and len(explanation) > 200:
            story.append(Paragraph("DETAILED ANALYSIS", section_style))
            
            # Split into paragraphs
            for para in explanation.split('\n'):
                if para.strip():
                    story.append(Paragraph(para.strip(), styles['Normal']))
                    story.append(Spacer(1, 0.05*inch))
            
            story.append(Spacer(1, 0.2*inch))
        
        # VirusTotal Report
        story.append(Paragraph("VIRUSTOTAL ANALYSIS", section_style))
        if vt_data and 'error' not in vt_data:
            stats = vt_data.get("last_analysis_stats", {})
            if not stats:
                stats = vt_data.get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            
            if stats:
                vt_table_data = [['Category', 'Count']] + [[k.capitalize(), str(v)] for k, v in stats.items()]
                vt_table = Table(vt_table_data, colWidths=[3*inch, 3.5*inch])
                vt_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4cc9f0')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('PADDING', (0, 0), (-1, -1), 8),
                ]))
                story.append(vt_table)
            else:
                story.append(Paragraph("No detection statistics available", styles['Normal']))
        else:
            story.append(Paragraph("No VirusTotal data available", styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
        
        # Shodan Report
        story.append(Paragraph("SHODAN NETWORK INTELLIGENCE", section_style))
        if shodan_data and 'error' not in shodan_data:
            shodan_rows = [
                ['IP Address', shodan_data.get("ip_str", shodan_data.get("ip", "N/A"))],
                ['Organization', shodan_data.get("org", "N/A")],
                ['ISP', shodan_data.get("isp", "N/A")],
                ['Country', shodan_data.get("country_name", "N/A")]
            ]
            
            ports = shodan_data.get("ports", [])
            if not ports and "data" in shodan_data:
                ports = [item.get("port") for item in shodan_data["data"] if "port" in item]
            
            shodan_rows.append(['Open Ports', ", ".join(map(str, ports[:20])) if ports else "None"])
            
            shodan_table = Table(shodan_rows, colWidths=[2*inch, 4.5*inch])
            shodan_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(shodan_table)
        else:
            story.append(Paragraph("No Shodan data available", styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
        
        # OTX Report
        story.append(Paragraph("ALIENVAULT OTX THREAT INTELLIGENCE", section_style))
        if otx_data and 'error' not in otx_data:
            otx_rows = [
                ['Threat Score', f"{otx_data.get('threat_score', 0)}/100"],
                ['Classification', otx_data.get('classification', 'Unknown')],
                ['Threat Pulses', str(otx_data.get('source_count', 0))]
            ]
            
            details = otx_data.get('details', {})
            if details:
                otx_rows.append(['Risk Level', details.get('risk_level', 'Unknown')])
                
                malware = details.get('malware_families', [])
                if malware:
                    otx_rows.append(['Malware Families', ", ".join(malware[:5])])
                
                tags = details.get('top_tags', [])
                if tags:
                    otx_rows.append(['Top Tags', ", ".join(tags[:10])])
            
            otx_table = Table(otx_rows, colWidths=[2*inch, 4.5*inch])
            otx_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(otx_table)
        else:
            story.append(Paragraph("No OTX data available", styles['Normal']))
        
        story.append(Spacer(1, 0.2*inch))

        # AbuseIPDB Report
        story.append(Paragraph("ABUSEIPDB REPUTATION", section_style))
        if abuseipdb_data and 'error' not in abuseipdb_data:
            abuse_table_data = [
                ['Threat Score', f"{abuseipdb_data.get('threat_score', 0)}/100"],
                ['Abuse Confidence', f"{abuseipdb_data.get('abuse_confidence_score', 0)}%"],
                ['Classification', abuseipdb_data.get('classification', 'Unknown')],
                ['Total Reports', str(abuseipdb_data.get('total_reports', 0))],
                ['Distinct Reporters', str(abuseipdb_data.get('details', {}).get('distinct_reporters', abuseipdb_data.get('total_reports', 0)))],
                ['Last Reported', abuseipdb_data.get('last_reported', 'N/A')],
                ['Country', abuseipdb_data.get('country_code', 'Unknown')],
                ['Usage Type', abuseipdb_data.get('usage_type', 'Unknown')],
                ['ISP', abuseipdb_data.get('isp', 'Unknown')],
                ['Domain', abuseipdb_data.get('domain', 'Unknown')]
            ]

            abuse_table = Table(abuse_table_data, colWidths=[2.2*inch, 4.3*inch])
            abuse_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(abuse_table)

            categories = abuseipdb_data.get('categories') or abuseipdb_data.get('details', {}).get('categories', {})
            categories = categories if isinstance(categories, dict) else {}
            if categories:
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph("Category Breakdown", styles['Heading3']))
                for name, count in categories.items():
                    story.append(Paragraph(f"• {name}: {count} reports", styles['Normal']))
        else:
            story.append(Paragraph("No AbuseIPDB data available or lookup not applicable", styles['Normal']))
        story.append(Spacer(1, 0.2*inch))

        # Google Search Intelligence
        story.append(Paragraph("OPEN-SOURCE INTELLIGENCE (GOOGLE)", section_style))
        if google_formatted:
            google_table_data = [
                ['Search Results', str(google_formatted.get('result_count', 0))],
                ['Threat Score', f"{google_formatted.get('threat_score', 0)}/100"],
                ['Classification', google_formatted.get('classification', 'Unknown')],
                ['Risk Level', google_formatted.get('risk_level', 'Unknown')],
                ['Unique Domains', str(google_formatted.get('total_domains', 0))]
            ]

            google_table = Table(google_table_data, colWidths=[2.2*inch, 4.3*inch])
            google_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(google_table)

            threat_indicators = google_formatted.get('threat_indicators', [])
            if threat_indicators:
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph("Threat Indicators", styles['Heading3']))
                story.append(Spacer(1, 0.05*inch))
                for indicator in threat_indicators:
                    story.append(Paragraph(f"• {indicator}", styles['Normal']))

            safe_indicators = google_formatted.get('safe_indicators', [])
            if safe_indicators:
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph("Safe Indicators", styles['Heading3']))
                story.append(Spacer(1, 0.05*inch))
                for indicator in safe_indicators:
                    story.append(Paragraph(f"• {indicator}", styles['Normal']))

            top_results = google_formatted.get('top_results', [])
            if top_results:
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph("Top Search Results", styles['Heading3']))
                story.append(Spacer(1, 0.05*inch))
                for entry in top_results:
                    story.append(Paragraph(f"<b>{entry.get('title', 'No title')}</b>", styles['Normal']))
                    story.append(Paragraph(entry.get('url', '#'), styles['Normal']))
                    snippet = entry.get('snippet', 'No description')
                    story.append(Paragraph(snippet, styles['Normal']))
                    story.append(Spacer(1, 0.05*inch))
        else:
            story.append(Paragraph("No Google search intelligence available", styles['Normal']))

        # Footer
        story.append(Spacer(1, 0.5*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        story.append(Paragraph("━" * 100, footer_style))
        story.append(Paragraph("This report was generated automatically by the Threat Intelligence Scanner", footer_style))
        story.append(Paragraph(f"Report ID: {str(result.id)}", footer_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"threat_report_{result.input_value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
    
    else:
        return jsonify({
            "error": "Invalid format. Supported formats: json, csv, excel, pdf"
        }), 400
# ---- HISTORY (Protected) ----
@main_bp.route("/history")
@login_required
def history():
    """View search history"""
    user_id = session.get("user_id")
    try:
        user = User.objects.get(id=user_id)
        results = IOCResult.objects(user_id=user).order_by('-timestamp')
    except:
        results = []
    
    return render_template("history.html", results=results)


# ---- LOGIN PAGE ----
@main_bp.route("/login", methods=["GET", "POST"])
def login():
    if 'user_id' in session:
        return redirect(url_for('main.index'))
    
    form = LoginForm()
    if form.validate_on_submit():
        try:
            email_input = form.email.data.strip()
            user = User.objects(email__iexact=email_input).first()
            if user and check_password_hash(user.password, form.password.data):
                session["user_id"] = str(user.id)
                session["username"] = user.email
                flash("Login successful!", "success")
                
                # Redirect to original page if stored
                next_page = session.pop('next', None)
                return redirect(next_page or url_for('main.index'))
            else:
                flash("Invalid email or password", "danger")
        except Exception as exc:
            logger.exception("Login error for email %s: %s", form.email.data, exc)
            flash("Invalid email or password", "danger")
    
    return render_template("login.html", form=form)


# ---- SIGNUP PAGE ----
@main_bp.route("/signup", methods=["GET", "POST"])
def signup():
    if 'user_id' in session:
        return redirect(url_for('main.index'))
    
    form = SignupForm()
    if form.validate_on_submit():
        # ✅ Email validation is handled by form validator
        # No need to check here, form.validate_email() does it
        email_input = form.email.data.strip()
        existing_user = User.objects(email__iexact=email_input).first()
        if existing_user:
            flash("That email is already registered. Please log in or use a different email.", "danger")
            return render_template("signup.html", form=form)

        hashed_pw = generate_password_hash(form.password.data, method='pbkdf2:sha256')
        new_user = User(email=email_input.lower(), password=hashed_pw)
        try:
            new_user.save()
            flash("Account created successfully! Please login.", "success")
            return redirect(url_for("main.login"))
        except Exception as exc:
            logger.exception("Signup error for email %s: %s", email_input, exc)
            flash("We couldn't create your account right now. Please try again.", "danger")
            return render_template("signup.html", form=form)
    
    return render_template("signup.html", form=form)

# ---- DASHBOARD (Protected) ----
@main_bp.route("/dashboard")
@login_required
def dashboard():
    """
    Threat Correlation Dashboard
    """
    try:
        # Get all dashboard data
        stats = get_dashboard_stats()
        top_ips, top_ip_scope = get_top_malicious_ips(limit=10)
        top_domains, top_domain_scope = get_top_malicious_domains(limit=10)
        geo_data = get_geolocation_data()
        timeline = get_threat_timeline(days=30)
        recent_threats = get_recent_threats(limit=20)
        classification_breakdown = get_classification_breakdown()
        top_tags = get_top_threat_tags(limit=15)
        
        # Debug logging
        logger.info(f"Dashboard data - Stats: {stats is not None}, Top IPs: {len(top_ips)}, Geo Data: {len(geo_data)}")
        logger.info(f"Top IPs data: {top_ips}")
        logger.info(f"Geo data: {geo_data}")
        
        return render_template(
            'dashboard.html',
            stats=stats,
            top_ips=top_ips,
            top_domains=top_domains,
            top_ip_scope=top_ip_scope,
            top_domain_scope=top_domain_scope,
            geo_data=geo_data,
            timeline=timeline,
            recent_threats=recent_threats,
            classification_breakdown=classification_breakdown,
            top_tags=top_tags
        )
    except Exception as e:
        logger.error(f"Dashboard error: {e}", exc_info=True)
        flash("Error loading dashboard", "danger")
        return redirect(url_for('main.index'))


# ---- API ENDPOINT: Real-time Threat Feed ----
@main_bp.route("/api/threats/recent")
@login_required
def api_recent_threats():
    """
    API endpoint for real-time threat feed updates
    """
    try:
        limit = request.args.get('limit', 20, type=int)
        recent = get_recent_threats(limit=limit)
        return jsonify(recent)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
# ---- LOGOUT ----
@main_bp.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully", "info")
    return redirect(url_for("main.landing"))